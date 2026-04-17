import json
import os
import openpyxl
from datetime import datetime
from funds import get_funds, get_funds_bond
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill


PERIOD_KEYS = [
    'netAssetValueRestoredGrowthRateRecentWeek',
    'netAssetValueRestoredGrowthRateRecentMonth',
    'netAssetValueRestoredGrowthRateRecentThreeMonth',
    'netAssetValueRestoredGrowthRateRecentSixMonth',
    'netAssetValueRestoredGrowthRateRecentOneYear',
    'netAssetValueRestoredGrowthRateRecentTwoYear',
    'netAssetValueRestoredGrowthRateRecentThreeYear',
    'netAssetValueRestoredGrowthRateRecentFiveYear',
]

PERIOD_LABELS = ['近一周', '1周~1月', '1月~3月', '3月~6月', '6月~1年', '1年~2年', '2年~3年', '3年~5年']
TOT_METRIC = 8


def parse_percent(val):
    return float(val[:-1])


def process_item(json_file):
    code = json_file['fundCode']
    name = json_file['name']
    total_asset = json_file['fund_manager_total_asset']

    week = round(parse_percent(json_file[PERIOD_KEYS[0]]), 2)
    increments = [week]

    prev = week + 100
    for key in PERIOD_KEYS[1:]:
        try:
            cur = parse_percent(json_file[key]) + 100
            increments.append(round(cur / prev * 100 - 100, 2))
            prev = cur
        except Exception:
            increments.append('--')

    return (code, name, *increments, total_asset, json_file['managerTrigger'])


def write_excel_section(worksheet, start_row, funds_config, compare_index, tmp_data,
                        highlight_red, highlight_green, fill_red, fill_green, fill_purple):
    hold_index = funds_config.get('hold_index', [])
    fund_list = funds_config['fund']

    title1 = ['', '', '']
    for index in compare_index:
        title1.append(tmp_data[index][1])
        title1.extend([''] * TOT_METRIC)
    title2 = ['代码', '名字', '基金经理管理规模']
    for _ in compare_index:
        title2.extend(PERIOD_LABELS + [''])

    for i, val in enumerate(title1):
        worksheet.cell(start_row, i + 1, val)
    for i, val in enumerate(title2):
        worksheet.cell(start_row + 1, i + 1, val)

    change_manager = []
    for ind, item in enumerate(fund_list):
        row = start_row + 2 + ind
        fund_id = worksheet.cell(row, 1, item)
        if item in hold_index:
            fund_id.fill = fill_purple
        worksheet.cell(row, 2, tmp_data[item][1])
        asset = worksheet.cell(row, 3, tmp_data[item][-2])
        asset_val = float(tmp_data[item][-2])
        if asset_val <= 100:
            asset.fill = fill_red
        elif asset_val > 300:
            asset.fill = fill_green

        cnt = 4
        for index in compare_index:
            for i in range(TOT_METRIC):
                try:
                    value = tmp_data[item][i + 2] - tmp_data[index][i + 2]
                    c = worksheet.cell(row, cnt, value)
                    if i < len(highlight_green) and value < highlight_green[i]:
                        c.fill = fill_green
                    if i < len(highlight_red) and value > highlight_red[i]:
                        c.fill = fill_red
                except Exception:
                    worksheet.cell(row, cnt, '--')
                cnt += 1
            cnt += 1

        if '又' not in tmp_data[item][-1]:
            try:
                if int(tmp_data[item][-1][:-1]) < 20:
                    change_manager.append((tmp_data[item][0], tmp_data[item][1]))
            except ValueError:
                pass

    return change_manager, len(fund_list)


def write_excel_signals(workbook, signals, tmp_data):
    ws = workbook.create_sheet(title="技术信号")
    headers = [
        '代码', '名字', '最新净值', '日期',
        'MA信号', 'MA得分', 'RSI值', 'RSI信号', 'RSI得分', 'MACD信号', 'MACD得分',
        'KDJ(J)', 'KDJ交叉', 'KDJ交叉得分', 'J值信号', 'J值得分',
        '布林%B', '布林信号', '布林得分',
        'ADX', '市场状态', 'ATR%',
        '百分位%', '百分位得分',
        '趋势60', '趋势120', '365日回撤%', '评分', '综合建议',
    ]
    for i, h in enumerate(headers):
        ws.cell(1, i + 1, h)

    green = Color(indexed=3, tint=0.5)
    fill_green = PatternFill('solid', fgColor=green)
    red = Color(indexed=2, tint=0.5)
    fill_red = PatternFill('solid', fgColor=red)

    row = 2
    for code, sig in signals.items():
        name = tmp_data[code][1] if code in tmp_data else code
        col = 1
        for val in [
            code, name, sig.get('latest_nav', ''), sig.get('latest_date', ''),
            sig['ma_signal'], sig.get('ma_score', 0),
            sig.get('rsi_value', ''), sig['rsi_signal'], sig.get('rsi_score', 0),
            sig['macd_signal'], sig.get('macd_score', 0),
            sig.get('j_value', ''), sig.get('kdj_cross_signal', ''), sig.get('kdj_cross_score', 0),
            sig.get('j_signal', ''), sig.get('j_score', 0),
            sig.get('pct_b', ''), sig.get('boll_signal', ''), sig.get('boll_score', 0),
            sig.get('adx_value', ''), sig.get('market_state', ''), sig.get('atr_pct', ''),
            sig.get('nav_percentile', ''), sig.get('percentile_score', ''),
            sig.get('trend_60', ''), sig.get('trend_120', ''),
            sig.get('max_drawdown', ''), sig.get('score', ''),
        ]:
            ws.cell(row, col, val)
            col += 1
        overall_cell = ws.cell(row, col, sig['overall'])
        if sig['overall'] == '买入':
            overall_cell.fill = fill_red
        elif sig['overall'] == '卖出':
            overall_cell.fill = fill_green
        row += 1

    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['D'].width = 14


def compute_excess_table(fund_list, compare_index, tmp_data, highlight_red, highlight_green, hold_index=None):
    hold_index = hold_index or []
    rows = []
    for item in fund_list:
        row = {
            'code': item,
            'name': tmp_data[item][1],
            'total_asset': tmp_data[item][-2],
            'is_held': item in hold_index,
            'manager_trigger': tmp_data[item][-1],
            'benchmarks': [],
        }
        asset_val = float(tmp_data[item][-2])
        row['asset_class'] = 'red' if asset_val <= 100 else ('green' if asset_val > 300 else '')

        for index in compare_index:
            cells = []
            for i in range(TOT_METRIC):
                try:
                    value = round(tmp_data[item][i + 2] - tmp_data[index][i + 2], 2)
                    css = ''
                    if i < len(highlight_green) and value < highlight_green[i]:
                        css = 'green'
                    if i < len(highlight_red) and value > highlight_red[i]:
                        css = 'red'
                    cells.append((f'{value:.2f}', css))
                except Exception:
                    cells.append(('--', ''))
            row['benchmarks'].append(cells)
        rows.append(row)
    return rows


# ============================================================
# SVG 走势图（250日 + MA5/MA20/MA60 + 布林带）
# ============================================================

def render_sparkline_svg(sig, width=650, height=140):
    """生成带布林带、MA60、买卖标记的大走势图"""
    navs = sig.get('recent_navs', [])
    ma5 = sig.get('recent_ma5', [])
    ma20 = sig.get('recent_ma20', [])
    ma60 = sig.get('recent_ma60', [])
    boll_upper = sig.get('recent_boll_upper', [])
    boll_lower = sig.get('recent_boll_lower', [])
    dates = sig.get('recent_dates', [])
    buy_markers = sig.get('buy_markers', [])
    sell_markers = sig.get('sell_markers', [])
    force_sell_markers = sig.get('force_sell_markers', [])

    valid_navs = [v for v in navs if v is not None and v != 'null']
    if len(valid_navs) < 2:
        return ''

    # 收集所有有效值来确定Y轴范围
    all_vals = list(valid_navs)
    for series in [boll_upper, boll_lower]:
        for v in series:
            if v is not None and v != 'null':
                all_vals.append(v)

    pad = 8
    chart_w = width - 2 * pad
    chart_h = height - 2 * pad
    nav_min = min(all_vals)
    nav_max = max(all_vals)
    nav_range = nav_max - nav_min
    if nav_range == 0:
        nav_range = 1

    n = len(navs)

    def to_xy(i, v):
        x = pad + i / (n - 1) * chart_w if n > 1 else pad
        y = pad + chart_h - (v - nav_min) / nav_range * chart_h
        return x, y

    def polyline_str(series):
        pts = []
        for i, v in enumerate(series):
            if v is not None and v != 'null':
                x, y = to_xy(i, v)
                pts.append(f'{x:.1f},{y:.1f}')
        return ' '.join(pts)

    svg = f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg">'

    # 布林带填充区域
    upper_pts = []
    lower_pts = []
    for i in range(n):
        u = boll_upper[i] if i < len(boll_upper) else None
        l = boll_lower[i] if i < len(boll_lower) else None
        if u is not None and u != 'null' and l is not None and l != 'null':
            ux, uy = to_xy(i, u)
            lx, ly = to_xy(i, l)
            upper_pts.append(f'{ux:.1f},{uy:.1f}')
            lower_pts.append(f'{lx:.1f},{ly:.1f}')
    if upper_pts and lower_pts:
        polygon = ' '.join(upper_pts) + ' ' + ' '.join(reversed(lower_pts))
        svg += f'<polygon points="{polygon}" fill="#e6f0fa" fill-opacity="0.5" stroke="none" />'

    # 均线和净值
    svg += f'<polyline points="{polyline_str(navs)}" fill="none" stroke="#1890ff" stroke-width="1.5" />'
    ma5_str = polyline_str(ma5)
    if ma5_str:
        svg += f'<polyline points="{ma5_str}" fill="none" stroke="#faad14" stroke-width="1" stroke-dasharray="3,2" />'
    ma20_str = polyline_str(ma20)
    if ma20_str:
        svg += f'<polyline points="{ma20_str}" fill="none" stroke="#f5222d" stroke-width="1" stroke-dasharray="5,3" />'
    ma60_str = polyline_str(ma60)
    if ma60_str:
        svg += f'<polyline points="{ma60_str}" fill="none" stroke="#722ed1" stroke-width="1" stroke-dasharray="8,4" />'

    # X轴月份刻度（根据数据跨度自动稀疏，避免重叠）
    all_months = []
    seen_months = set()
    for i, d in enumerate(dates):
        if not d or d == 'null':
            continue
        month_key = d[:7]
        if month_key not in seen_months:
            seen_months.add(month_key)
            all_months.append((i, d))

    # 根据总月数决定间隔：<=24月每月显示，<=60月每3月，>60月每6月
    total_months = len(all_months)
    if total_months <= 24:
        step = 1
    elif total_months <= 60:
        step = 3
    else:
        step = 6

    for idx, (i, d) in enumerate(all_months):
        x, _ = to_xy(i, nav_min)
        # 细网格线每月都画
        svg += f'<line x1="{x:.1f}" y1="{pad}" x2="{x:.1f}" y2="{height - pad}" stroke="#e8e8e8" stroke-width="0.5" />'
        # 文字标签按间隔显示
        if idx % step == 0:
            month_num = d[5:7]
            label = f'{d[:4]}' if month_num == '01' else f'{month_num}月'
            svg += f'<text x="{x:.1f}" y="{height - 1}" text-anchor="middle" font-size="9" fill="#999">{label}</text>'

    # 买入标记（红色向上三角）
    for ci, nav_val in buy_markers:
        if nav_val is not None and nav_val != 'null':
            x, y = to_xy(ci, nav_val)
            svg += f'<polygon points="{x:.1f},{y + 3:.1f} {x - 4:.1f},{y + 11:.1f} {x + 4:.1f},{y + 11:.1f}" fill="#f5222d" opacity="0.85" />'

    # 卖出标记（绿色向下三角）
    for ci, nav_val in sell_markers:
        if nav_val is not None and nav_val != 'null':
            x, y = to_xy(ci, nav_val)
            svg += f'<polygon points="{x:.1f},{y - 3:.1f} {x - 4:.1f},{y - 11:.1f} {x + 4:.1f},{y - 11:.1f}" fill="#52c41a" opacity="0.85" />'

    # 止盈标记（紫色向下三角）
    for ci, nav_val in force_sell_markers:
        if nav_val is not None and nav_val != 'null':
            x, y = to_xy(ci, nav_val)
            svg += f'<polygon points="{x:.1f},{y - 3:.1f} {x - 4:.1f},{y - 11:.1f} {x + 4:.1f},{y - 11:.1f}" fill="#722ed1" opacity="0.85" />'

    svg += '</svg>'
    return svg


def render_rsi_bar(rsi_value):
    if rsi_value is None:
        return '<span class="rsi-na">N/A</span>'
    pct = max(0, min(100, rsi_value))
    if pct > 70:
        color = '#cf1322'
        label = '超买'
    elif pct < 30:
        color = '#389e0d'
        label = '超卖'
    else:
        color = '#1890ff'
        label = '中性'
    return (
        f'<div class="rsi-bar">'
        f'<div class="rsi-fill" style="width:{pct}%;background:{color}"></div>'
        f'<span class="rsi-label">{rsi_value} ({label})</span>'
        f'</div>'
    )


def render_signal_badge(signal, score=None):
    cls = 'badge-buy' if signal == '买入' else ('badge-sell' if signal == '卖出' else 'badge-hold')
    score_str = f' ({score:+g})' if score is not None and score != 0 else ''
    return f'<span class="signal-badge {cls}">{signal}{score_str}</span>'


def render_trend_badge(trend):
    if trend == '多头':
        return '<span class="trend-bull">多头</span>'
    elif trend == '空头':
        return '<span class="trend-bear">空头</span>'
    return '<span class="trend-neutral">未知</span>'


def render_percentile_bar(percentile):
    if percentile is None:
        return '<span class="rsi-na">N/A</span>'
    pct = max(0, min(100, percentile))
    if pct < 10:
        color = '#389e0d'
        label = '极低'
    elif pct < 20:
        color = '#52c41a'
        label = '偏低'
    elif pct > 90:
        color = '#cf1322'
        label = '极高'
    elif pct > 80:
        color = '#f5222d'
        label = '偏高'
    else:
        color = '#1890ff'
        label = '中性'
    return (
        f'<div class="rsi-bar">'
        f'<div class="rsi-fill" style="width:{pct}%;background:{color}"></div>'
        f'<span class="rsi-label">{percentile:.1f}% ({label})</span>'
        f'</div>'
    )


def render_adx_bar(adx_value, market_state):
    if adx_value is None:
        return '<span class="rsi-na">N/A</span>'
    pct = max(0, min(100, adx_value))
    if adx_value >= 25:
        color = '#1890ff'
        label = '趋势'
    else:
        color = '#faad14'
        label = '震荡'
    return (
        f'<div class="rsi-bar">'
        f'<div class="rsi-fill" style="width:{pct}%;background:{color}"></div>'
        f'<span class="rsi-label">{adx_value:.1f} ({label})</span>'
        f'</div>'
    )


def render_atr_display(atr_pct):
    if atr_pct is None:
        return '<span class="rsi-na">N/A</span>'
    if atr_pct > 2.0:
        cls = 'dd-high'
    elif atr_pct > 1.0:
        cls = 'dd-mid'
    else:
        cls = 'dd-low'
    return f'<span class="{cls}">{atr_pct:.3f}%</span>'


def render_score_bar(score, max_score=7):
    """评分条：动态范围"""
    pct = (score + max_score) / (2 * max_score) * 100
    pct = max(0, min(100, pct))
    if score >= 4:
        color = '#f5222d'
    elif score <= -4:
        color = '#52c41a'
    elif score > 0:
        color = '#ff7875'
    elif score < 0:
        color = '#95de64'
    else:
        color = '#8c8c8c'
    return (
        f'<div class="score-bar">'
        f'<div class="score-fill" style="width:{pct}%;background:{color}"></div>'
        f'<span class="score-label">{score:+.1f}</span>'
        f'</div>'
    )


def render_drawdown(dd):
    if dd is None:
        return '--'
    if dd > 15:
        cls = 'dd-high'
    elif dd > 8:
        cls = 'dd-mid'
    else:
        cls = 'dd-low'
    return f'<span class="{cls}">{dd:.1f}%</span>'


def generate_html(tmp_data, equity_config, bond_config, change_manager, signals=None):
    equity_compare = equity_config['compare_index']
    bond_compare = bond_config['compare_index']

    equity_highlight_red = [1.5, 3, 5, 7, 12, 20, 20, 25]
    equity_highlight_green = [-1.5, -3, -4.5, -6, -10, -15, -15, -20]
    bond_highlight_red = [0.04, 0.1, 0.3, 0.4, 0.75, 1.5, 1.5]
    bond_highlight_green = [-0.03, 0, -0.1, -0.15, -0.25, -0.5, -0.5]

    equity_rows = compute_excess_table(
        equity_config['fund'], equity_compare, tmp_data,
        equity_highlight_red, equity_highlight_green, equity_config.get('hold_index', []))
    bond_rows = compute_excess_table(
        bond_config['fund'], bond_compare, tmp_data,
        bond_highlight_red, bond_highlight_green)

    today = datetime.now().strftime('%Y-%m-%d %H:%M')

    def render_table(title, compare_index, rows):
        benchmark_names = [tmp_data[idx][1] for idx in compare_index]
        n_benchmarks = len(compare_index)
        header1 = '<tr><th></th><th></th><th></th>'
        for name in benchmark_names:
            header1 += f'<th colspan="{TOT_METRIC}" class="benchmark-header">{name}</th>'
        header1 += '</tr>'

        header2 = '<tr><th>代码</th><th>名字</th><th>管理规模(亿)</th>'
        for _ in range(n_benchmarks):
            for label in PERIOD_LABELS:
                header2 += f'<th>{label}</th>'
        header2 += '</tr>'

        body = ''
        for row in rows:
            code_class = ' class="held"' if row['is_held'] else ''
            asset_cls = f' class="{row["asset_class"]}"' if row['asset_class'] else ''
            body += '<tr>'
            body += f'<td{code_class}>{row["code"]}</td>'
            body += f'<td>{row["name"]}</td>'
            body += f'<td{asset_cls}>{row["total_asset"]}</td>'
            for cells in row['benchmarks']:
                for val, css in cells:
                    cls = f' class="{css}"' if css else ''
                    body += f'<td{cls}>{val}</td>'
            body += '</tr>\n'

        return f'''
        <h2>{title}</h2>
        <div class="table-wrapper">
        <table>
            <thead>{header1}{header2}</thead>
            <tbody>{body}</tbody>
        </table>
        </div>'''

    manager_html = ''
    if change_manager:
        items = ''.join(f'<li>{code} - {name}</li>' for code, name in change_manager)
        manager_html = f'<div class="alert">近20天内以下基金经理发生变更：<ul>{items}</ul></div>'
    else:
        manager_html = '<div class="ok">近20天内列表中基金经理没有变更。</div>'

    # 技术分析板块
    signal_html = ''
    if signals:
        all_fund_codes = equity_config['fund'] + bond_config['fund']
        signal_rows = ''
        for code in all_fund_codes:
            if code not in signals:
                continue
            sig = signals[code]
            name = tmp_data[code][1] if code in tmp_data else code

            sparkline = render_sparkline_svg(sig)

            j_val = sig.get('j_value', '')
            j_display = f'{j_val:.1f}' if isinstance(j_val, (int, float)) else '--'
            pct_b = sig.get('pct_b', '')
            pct_b_display = f'{pct_b:.2f}' if isinstance(pct_b, (int, float)) else '--'

            # 净值列：如果有估算数据，追加显示
            nav_display = str(sig.get('latest_nav', ''))
            if sig.get('estimated') and sig.get('latest_nav'):
                nav_display += f'<br><small style="color:#1890ff">估值</small>'

            # 建议列：追加估算时间
            gztime = sig.get('gztime', '')
            gztime_display = ''
            if gztime:
                # "2026-04-14 13:52" → "04-14 13:52"
                gztime_short = gztime[5:] if len(gztime) > 5 else gztime
                gztime_display = f'<br><small>估算: {gztime_short}</small>'

            ma_s = sig.get('ma_score', 0)
            rsi_s = sig.get('rsi_score', 0)
            macd_s = sig.get('macd_score', 0)
            kdj_cross_s = sig.get('kdj_cross_score', 0)
            j_s = sig.get('j_score', 0)
            boll_s = sig.get('boll_score', 0)

            force_note = '<br><small style="color:#722ed1">触发止盈</small>' if sig.get('is_force_sell') else ''
            if sig.get('is_force_sell'):
                overall_badge = '<span class="signal-badge badge-force-sell">止盈信号</span>'
            else:
                overall_badge = render_signal_badge(sig['overall'])

            signal_rows += f'''<tr>
                <td>{code}</td>
                <td>{name}</td>
                <td>{nav_display}</td>
                <td class="chart-cell">{sparkline}</td>
                <td>{render_signal_badge(sig['ma_signal'], ma_s)}</td>
                <td>{render_rsi_bar(sig.get('rsi_value'))}<br><small>得分: {rsi_s:+g}</small></td>
                <td>{render_signal_badge(sig['macd_signal'], macd_s)}</td>
                <td>{render_signal_badge(sig.get('kdj_cross_signal', '持有'), kdj_cross_s)}<br>{render_signal_badge(sig.get('j_signal', '持有'), j_s)}<br><small>J={j_display}</small></td>
                <td>{render_signal_badge(sig.get('boll_signal', '持有'), boll_s)}<br><small>%B={pct_b_display}</small></td>
                <td>{render_adx_bar(sig.get('adx_value'), sig.get('market_state', ''))}</td>
                <td>{render_atr_display(sig.get('atr_pct'))}</td>
                <td>{render_percentile_bar(sig.get('nav_percentile'))}<br></td>
                <td>{render_score_bar(sig.get('score', 0))}{force_note}</td>
                <td>{overall_badge}{gztime_display}</td>
            </tr>\n'''

        signal_html = f'''
        <h2>技术分析 - 多维买卖信号</h2>
        <div class="legend">
            <span><svg width="20" height="10"><line x1="0" y1="5" x2="20" y2="5" stroke="#1890ff" stroke-width="1.5"/></svg> 净值</span>
            <span><svg width="20" height="10"><line x1="0" y1="5" x2="20" y2="5" stroke="#faad14" stroke-width="1" stroke-dasharray="3,2"/></svg> MA5</span>
            <span><svg width="20" height="10"><line x1="0" y1="5" x2="20" y2="5" stroke="#f5222d" stroke-width="1" stroke-dasharray="5,3"/></svg> MA20</span>
            <span><svg width="20" height="10"><line x1="0" y1="5" x2="20" y2="5" stroke="#722ed1" stroke-width="1" stroke-dasharray="8,4"/></svg> MA60</span>
            <span><svg width="20" height="10"><rect x="0" y="2" width="20" height="6" fill="#e6f0fa" opacity="0.7"/></svg> 布林带</span>
            <span><svg width="14" height="12"><polygon points="7,1 2,11 12,11" fill="#f5222d" opacity="0.85"/></svg> 买入信号</span>
            <span><svg width="14" height="12"><polygon points="7,11 2,1 12,1" fill="#52c41a" opacity="0.85"/></svg> 卖出信号</span>
            <span><svg width="14" height="12"><polygon points="7,11 2,1 12,1" fill="#722ed1" opacity="0.85"/></svg> 止盈信号</span>
        </div>
        <p class="signal-note">ADX&ge;25=趋势行情（MA/MACD权重&times;1.5），ADX&lt;25=震荡行情（RSI/布林权重&times;1.5），KDJ固定权重&times;1 | 综合建议: 评分 &ge;+4 买入, &le;-4 卖出 | 买卖信号经历史去重(变动需超3%)、百分位过滤(买&gt;90%/卖&lt;10%无效) | 净值涨幅&gt;25%触发<span style="color:#722ed1">止盈信号</span> | ATR/百分位仅展示</p>
        <div class="table-wrapper">
        <table class="signal-table">
            <thead>
                <tr>
                    <th>代码</th><th>名字</th><th>净值</th><th>近5年走势</th>
                    <th>MA(5/20)</th><th>RSI(14)</th><th>MACD</th><th>KDJ(交叉/J值)</th><th>布林带</th>
                    <th>ADX(14)</th><th>ATR%</th><th>百分位（近6年）</th>
                    <th>评分</th><th>建议</th>
                </tr>
            </thead>
            <tbody>{signal_rows}</tbody>
        </table>
        </div>'''

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>基金量化指标报告</title>
<style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif; background: #f5f7fa; color: #333; padding: 20px; }}
    h1 {{ text-align: center; margin: 20px 0 5px; font-size: 24px; }}
    .date {{ text-align: center; color: #888; margin-bottom: 20px; font-size: 14px; }}
    h2 {{ margin: 30px 0 10px; padding-left: 10px; border-left: 4px solid #1890ff; font-size: 18px; }}
    .table-wrapper {{ overflow-x: auto; margin-bottom: 20px; }}
    table {{ border-collapse: collapse; font-size: 13px; white-space: nowrap; min-width: 100%; }}
    th, td {{ border: 1px solid #d9d9d9; padding: 6px 10px; text-align: center; }}
    thead th {{ background: #fafafa; font-weight: 600; position: sticky; top: 0; }}
    .benchmark-header {{ background: #e6f7ff; color: #1890ff; }}
    tbody tr:hover {{ background: #f0f5ff; }}
    td.red {{ background: #fff1f0; color: #cf1322; }}
    td.green {{ background: #f6ffed; color: #389e0d; }}
    td.held {{ background: #f9f0ff; color: #722ed1; font-weight: 600; }}
    .alert {{ margin: 20px 0; padding: 12px 16px; background: #fff2e8; border: 1px solid #ffbb96; border-radius: 4px; color: #d4380d; }}
    .alert ul {{ margin: 8px 0 0 20px; }}
    .ok {{ margin: 20px 0; padding: 12px 16px; background: #f6ffed; border: 1px solid #b7eb8f; border-radius: 4px; color: #389e0d; }}

    /* 技术分析样式 */
    .signal-table td {{ vertical-align: middle; }}
    .chart-cell {{ padding: 4px !important; }}
    .signal-badge {{
        display: inline-block; padding: 2px 8px; border-radius: 10px;
        font-size: 11px; font-weight: 600; color: #fff; margin: 1px;
    }}
    .badge-buy {{ background: #f5222d; }}
    .badge-sell {{ background: #52c41a; }}
    .badge-hold {{ background: #8c8c8c; }}
    .badge-force-sell {{ background: #722ed1; }}
    .trend-bull {{ color: #cf1322; font-weight: 600; font-size: 12px; }}
    .trend-bear {{ color: #389e0d; font-weight: 600; font-size: 12px; }}
    .trend-neutral {{ color: #8c8c8c; font-size: 12px; }}
    .rsi-bar {{
        position: relative; width: 110px; height: 18px;
        background: #f0f0f0; border-radius: 9px; overflow: hidden;
        display: inline-block; vertical-align: middle;
    }}
    .rsi-fill {{ height: 100%; border-radius: 9px; }}
    .rsi-label {{
        position: absolute; top: 0; left: 0; right: 0;
        text-align: center; font-size: 11px; line-height: 18px;
        color: #333; font-weight: 500;
    }}
    .rsi-na {{ color: #bbb; font-size: 12px; }}
    .score-bar {{
        position: relative; width: 80px; height: 18px;
        background: #f0f0f0; border-radius: 9px; overflow: hidden;
        display: inline-block; vertical-align: middle;
    }}
    .score-fill {{ height: 100%; border-radius: 9px; }}
    .score-label {{
        position: absolute; top: 0; left: 0; right: 0;
        text-align: center; font-size: 11px; line-height: 18px;
        color: #333; font-weight: 600;
    }}
    .dd-high {{ color: #cf1322; font-weight: 600; }}
    .dd-mid {{ color: #fa8c16; font-weight: 500; }}
    .dd-low {{ color: #389e0d; }}
    .legend {{
        margin: 10px 0; display: flex; gap: 16px; align-items: center;
        font-size: 13px; color: #666; flex-wrap: wrap;
    }}
    .legend span {{ display: flex; align-items: center; gap: 4px; }}
    .signal-note {{ font-size: 12px; color: #999; margin: 4px 0 12px; }}
    small {{ color: #999; font-size: 11px; }}
</style>
</head>
<body>
<h1>基金量化指标报告</h1>
<p class="date">生成时间：{today}</p>
{manager_html}
{render_table("股票型基金（中高风险）", equity_compare, equity_rows)}
{render_table("债券型基金（中低风险）", bond_compare, bond_rows)}
{signal_html}
</body>
</html>'''

    with open('fund_report.html', 'w', encoding='utf-8') as f:
        f.write(html)


if __name__ == '__main__':
    green = Color(indexed=3, tint=0.5)
    fill_green = PatternFill('solid', fgColor=green)
    red = Color(indexed=2, tint=0.5)
    fill_red = PatternFill('solid', fgColor=red)
    purple = Color(indexed=20, tint=0.3)
    fill_purple = PatternFill('solid', fgColor=purple)

    tmp_data = {}
    with open('data/temp.json', encoding='utf8') as file:
        for line in file:
            data = json.loads(line)
            item = process_item(data)
            tmp_data[item[0]] = item

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "fund"

    # 股票基金
    equity_config = get_funds()
    equity_highlight_red = [1.5, 3, 5, 7, 12, 20, 20, 25]
    equity_highlight_green = [-1.5, -3, -4.5, -6, -10, -15, -15, -20]
    change_manager, equity_count = write_excel_section(
        worksheet, 1, equity_config, equity_config['compare_index'], tmp_data,
        equity_highlight_red, equity_highlight_green, fill_red, fill_green, fill_purple)

    # 债券基金
    bond_config = get_funds_bond()
    bond_highlight_red = [0.04, 0.1, 0.3, 0.4, 0.75, 1.5, 1.5]
    bond_highlight_green = [-0.03, 0, -0.1, -0.15, -0.25, -0.5, -0.5]
    bond_start = 1 + equity_count + 5
    bond_change, _ = write_excel_section(
        worksheet, bond_start, bond_config, bond_config['compare_index'], tmp_data,
        bond_highlight_red, bond_highlight_green, fill_red, fill_green, fill_purple)
    change_manager.extend(bond_change)

    # 加载技术信号
    signals = None
    signals_file = 'data/signals.json'
    if os.path.exists(signals_file):
        with open(signals_file, encoding='utf-8') as f:
            signals = json.load(f)
        write_excel_signals(workbook, signals, tmp_data)

    worksheet.column_dimensions['B'].width = 32
    worksheet.column_dimensions['C'].width = 20
    workbook.save(filename='fund.xlsx')

    # 生成 HTML 报告
    generate_html(tmp_data, equity_config, bond_config, change_manager, signals)

    print('*' * 30, 'result', '*' * 30)
    print('处理完毕，文件 fund.xlsx 和 fund_report.html 已输出。')
    if signals:
        buy_count = sum(1 for s in signals.values() if s['overall'] == '买入')
        sell_count = sum(1 for s in signals.values() if s['overall'] == '卖出')
        hold_count = sum(1 for s in signals.values() if s['overall'] == '持有')
        print(f'技术分析信号：买入 {buy_count} 只，卖出 {sell_count} 只，持有 {hold_count} 只')
    if not change_manager:
        print('经检查，近20天内列表中基金的经理没有更换。')
    else:
        print('经检查，近20天以下基金的经理发生了人事变动：')
        for code, name in change_manager:
            print(f'  {code} - {name}')
